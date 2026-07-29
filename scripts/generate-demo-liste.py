from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import date, timedelta
import random
random.seed(7)

TN = [("PK01","Aydin","Elif"),("PK03","Kovac","Marek"),("PK06","Nowak","Zofia"),
      ("PK08","Haddad","Rania"),("PK09","Silva","Tomas"),("PK10","Petrov","Ivan"),
      ("PK11","Demir","Aylin"),("PK15","Okafor","Chidi"),("PK19","Rossi","Giulia"),
      ("PK20","Mensah","Kwame"),("PK23","Novak","Marta"),("PK24","Bauer","Lena")]
DAYL = {0:"M",1:"D",2:"M",3:"D",4:"F",5:"S",6:"S"}
MON = ["JANUAR","FEBRUAR","MÄRZ","APRIL","MAI","JUNI","JULI","AUGUST",
       "SEPTEMBER","OKTOBER","NOVEMBER","DEZEMBER"]
HOL = {2025:{"2025-01-01","2025-04-18","2025-04-21","2025-05-01","2025-05-29",
             "2025-06-09","2025-10-03","2025-10-31","2025-12-25","2025-12-26"},
       2026:{"2026-01-01","2026-04-03","2026-04-06","2026-05-01","2026-05-14",
             "2026-05-25","2026-09-20","2026-10-03","2026-10-31","2026-12-25","2026-12-26"}}
HDR = PatternFill("solid", fgColor="DDDDDD")

def weeks(year):
    d = date(year,1,1)
    d -= timedelta(days=d.weekday())
    out=[]
    while d.year <= year:
        days=[d+timedelta(days=i) for i in range(5)]
        if any(x.year==year for x in days): out.append(days)
        d += timedelta(days=7)
        if d.year > year and d.month==1: break
    return out

def build(ws, year):
    ws.append([]); r=1
    totals={(t[0],m):0 for t in TN for m in range(1,13)}
    for wk in weeks(year):
        # Wochen ueber Monatsgrenzen werden - wie in der echten Liste -
        # in zwei Bloecke geteilt.
        groups=[]
        cur=[wk[0]]
        for d in wk[1:]:
            if d.month==cur[-1].month: cur.append(d)
            else: groups.append(cur); cur=[d]
        groups.append(cur)
        offset=0
        for g in groups:
            r+=1
            head=[None]*20
            head[0]=MON[g[0].month-1] if g[0].month else None
            head[1]=f"KW{g[0].isocalendar()[1]}"; head[2]="TN-ID"
            head[3]="Nachname"; head[4]="Vorname"
            for i,d in enumerate(g):
                head[5+(offset+i)*2]=f"{DAYL[d.weekday()]} {d.day}.{d.month:02d}"
            head[19]="Anmerkungen"
            ws.append(head)
            for c in range(1,21):
                ws.cell(r,c).fill=HDR; ws.cell(r,c).font=Font(bold=True)
            r+=1
            vn=[None]*20
            for i in range(len(g)):
                vn[5+(offset+i)*2]="V"; vn[6+(offset+i)*2]="N"
            ws.append(vn)
            for tid,nn,vn_ in TN:
                r+=1
                row=[None]*21
                row[2]=tid; row[3]=nn; row[4]=vn_
                for i,d in enumerate(g):
                    iso=d.isoformat()
                    if iso in HOL[year] or d.year!=year: continue
                    if tid=="PK24" and d.month<4: continue
                    code=random.choices(["X","X","X","(x)","E","K","A","U",""],
                                        [50,20,10,5,5,3,2,3,2])[0]
                    row[5+(offset+i)*2]=code
                    row[6+(offset+i)*2]=code if code not in ("(x)",) else "X"
                    if code in ("X","(x)","E","K"): totals[(tid,d.month)]+=1
                if tid=="PK23" and g[0].month==3 and g[0].day>=28:
                    row[2]=2  # bewusster Datenfehler wie in der echten Datei
                if random.random()<0.05: row[19]="Nachweis fehlt"
                ws.append(row)
            offset+=len(g)
    return totals

wb=Workbook(); wb.remove(wb.active)
allt={}
for y in (2026,2025):
    ws=wb.create_sheet(str(y)); allt[y]=build(ws,y)

ws=wb.create_sheet("Regeln")
for row in [["Code","Bedeutung","Zaehlt als"],["X","Anwesend","Anwesend"],
            ["(x)","Anwesend, verspaetet/frueher","Anwesend"],
            ["E","Entschuldigt mit Nachweis","Anwesend"],
            ["K","Kulanztag","Anwesend"],["A","Abgemeldet ohne Nachweis","Fehltag"],
            ["U","Nicht abgemeldet","Fehltag"],["(leer)","Noch nicht erfasst","Offen"]]:
    ws.append(row)

ws=wb.create_sheet("Overall")
ws.append(["Jahr","TN-ID","Nachname","Vorname"]+MON)
for y in (2025,2026):
    for tid,nn,vnn in TN:
        ws.append([str(y),tid,nn,vnn]+[allt[y][(tid,m)] for m in range(1,13)])
wb.save("public/demo/Anwesenheitsliste_Demo.xlsx")
print("ok")
